import io
import time
import streamlit as st
import requests
import pandas as pd
import PyPDF2
import docx
from pathlib import Path
from evaluation import evaluate_summary

# Config
BACKEND       = "http://localhost:8000"
MAX_CHARS     = 30000
POLL_INTERVAL = 2

MODELS        = ["llama3.2", "phi3"]
SUMMARY_TYPES = ["comprehensive", "executive", "bullet_points"]
SUMMARY_LABELS = {
    "comprehensive": "📋 Comprehensive",
    "executive":     "👔 Executive",
    "bullet_points": "• Bullet Points",
}
DOC_CATEGORIES = ["general", "policy", "report", "user_guide", "financial", "correspondence"]
CATEGORY_LABELS = {
    "general":        "📄 General",
    "policy":         "📜 Policy",
    "report":         "📊 Report",
    "user_guide":     "📖 User Guide",
    "financial":      "💰 Financial",
    "correspondence": "✉️ Correspondence",
}

# Page setup
st.set_page_config(layout="wide", page_title="On-Prem AI Summarizer")

st.markdown("""
<style>
.main-header  {font-size:2.5rem; color:#1E88E5; text-align:center; margin-bottom:20px;}
.stButton>button {width:100%; border-radius:5px; height:50px; font-size:18px;}
.status-box   {padding:10px; border-radius:5px; background:#e8f5e9; border:1px solid #c8e6c9;}
.result-box   {background:#f8f9fa; border-left:4px solid #1E88E5; padding:1rem 1.2rem;
    font-size:0.92rem; line-height:1.7; border-radius:0 6px 6px 0; white-space:pre-wrap;}
.combo-header {background:#f5f5f5; padding:6px 12px; border-radius:4px;
    font-size:0.82rem; margin-bottom:4px;}
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">📄 Secure On-Prem Document Summarizer</div>', unsafe_allow_html=True)

def check_backend() -> bool:
    try:
        return requests.get(f"{BACKEND}/health", timeout=3).status_code == 200
    except Exception:
        return False

def get_model_status() -> list:
    try:
        resp = requests.get(f"{BACKEND}/models", timeout=3)
        if resp.status_code == 200:
            return resp.json().get("models", [])
    except Exception:
        pass
    return []

def extract_text(file) -> str | None:
    text = ""
    try:
        ext = Path(file.name).suffix.lower()
        if ext == ".pdf":
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                content = page.extract_text()
                if content:
                    text += content + "\n"
                if len(text) > MAX_CHARS:
                    text += "\n...(Text truncated for speed)..."
                    break
        elif ext == ".docx":
            doc = docx.Document(file)
            for para in doc.paragraphs:
                text += para.text + "\n"
                if len(text) > MAX_CHARS:
                    text += "\n...(Text truncated for speed)..."
                    break
        elif ext == ".txt":
            text = file.read().decode("utf-8")[:MAX_CHARS]
    except Exception as e:
        st.error(f"Error reading file: {e}")
        return None
    return text or None

def submit_job(text: str, model_name: str, summary_type: str) -> str | None:
    try:
        resp = requests.post(
            f"{BACKEND}/summarize",
            json={"text": text, "model_name": model_name, "summary_type": summary_type},
            timeout=10,
        )
        if resp.status_code == 200:
            return resp.json()["job_id"]
        st.error(f"Backend error: {resp.json().get('detail')}")
    except Exception as e:
        st.error(f"Connection Error: {e}")
    return None

def poll_until_done(job_id: str, label: str = "") -> dict | None:
    status_slot = st.empty()
    while True:
        try:
            resp = requests.get(f"{BACKEND}/job/{job_id}", timeout=5)
            if resp.status_code != 200:
                status_slot.error("Failed to poll job status.")
                return None
            data = resp.json()
            status = data["status"]
            if status == "completed":
                status_slot.empty()
                return data
            elif status == "failed":
                status_slot.error(f"❌ Job failed: {data.get('error')}")
                return None
            else:
                status_slot.info(f"⏳ {label} — **{status}**, checking in {POLL_INTERVAL}s...")
                time.sleep(POLL_INTERVAL)
        except Exception as e:
            status_slot.error(f"Polling error: {e}")
            return None

def poll_job_once(job_id: str) -> dict | None:
    try:
        resp = requests.get(f"{BACKEND}/job/{job_id}", timeout=5)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return None

def build_excel(results: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    HEADER_FILL  = PatternFill("solid", start_color="1565C0")
    GOOD_FILL    = PatternFill("solid", start_color="C8E6C9")
    BAD_FILL     = PatternFill("solid", start_color="FFCDD2")
    ALT_FILL     = PatternFill("solid", start_color="F5F5F5")
    WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
    HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BOLD_FONT    = Font(name="Arial", bold=True, size=10)
    NORMAL_FONT  = Font(name="Arial", size=10)
    thin         = Side(style="thin", color="BDBDBD")
    BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def hdr(cell):
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, HEADER_FONT, CENTER, BORDER

    def cell_style(cell, fill=WHITE_FILL, font=NORMAL_FONT, align=CENTER):
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, align, BORDER

    # Full Results
    ws1 = wb.active
    ws1.title = "Full Results"

    ws1.merge_cells("A1:I1")
    t = ws1["A1"]
    t.value = "DocSum — Model Comparison Matrix Results"
    t.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill  = HEADER_FILL
    t.alignment = CENTER
    ws1.row_dimensions[1].height = 28

    headers = ["Doc Category", "Model", "Summary Type", "Starts on Topic ✓",
    "Word Count", "Time (s)", "Readability Score", "Readability Label", "Summary Preview"]
    for col, h in enumerate(headers, 1):
        hdr(ws1.cell(row=2, column=col, value=h))
    ws1.row_dimensions[2].height = 22

    for i, r in enumerate(results):
        row  = i + 3
        fill = ALT_FILL if i % 2 == 0 else WHITE_FILL
        vals = [
            r.get("doc_category", ""),
            r.get("model", ""),
            r.get("summary_type", ""),
            "✅ Yes" if r.get("starts_on_topic") else "❌ No",
            r.get("word_count", 0),
            r.get("time_taken", 0),
            r.get("readability", 0),
            r.get("readability_label", ""),
            r.get("summary_preview", ""),
        ]
        for col, val in enumerate(vals, 1):
            c = ws1.cell(row=row, column=col, value=val)
            if col == 4:
                cell_style(c, fill=GOOD_FILL if r.get("starts_on_topic") else BAD_FILL, font=BOLD_FONT)
            else:
                cell_style(c, fill=fill, align=LEFT if col == 9 else CENTER)

    for col, w in enumerate([18,14,18,18,12,10,16,18,60], 1):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = "A3"

    # Pivot Matrix
    ws2 = wb.create_sheet("Comparison Matrix")
    df  = pd.DataFrame(results)

    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = "Speed & Quality Comparison — Model × Summary Type"
    t2.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t2.fill  = HEADER_FILL
    t2.alignment = CENTER
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(["Model","Summary Type","Avg Time (s)","Avg Words",
        "Avg Readability","Topic Start %","Overall Score*"], 1):
        hdr(ws2.cell(row=2, column=col, value=h))

    pivot_row = 3
    for model in MODELS:
        for stype in SUMMARY_TYPES:
            sub = df[(df["model"] == model) & (df["summary_type"] == stype)]
            if sub.empty:
                continue
            avg_time    = sub["time_taken"].mean()
            avg_words   = sub["word_count"].mean()
            avg_read    = sub["readability"].mean()
            topic_rate  = sub["starts_on_topic"].mean() * 100
            speed_score = max(0, 100 - avg_time)
            overall     = round(0.4*avg_read + 0.4*topic_rate + 0.2*speed_score, 1)

            vals = [model, stype, round(avg_time,1), round(avg_words,0),
                    round(avg_read,1), f"{topic_rate:.0f}%", overall]
            fill = ALT_FILL if pivot_row % 2 == 0 else WHITE_FILL
            for col, val in enumerate(vals, 1):
                c = ws2.cell(row=pivot_row, column=col, value=val)
                f = GOOD_FILL if (col==7 and overall>=60) else (BAD_FILL if (col==7 and overall<40) else fill)
                cell_style(c, fill=f, font=BOLD_FONT if col==7 else NORMAL_FONT)
            pivot_row += 1

    note_row = pivot_row + 1
    ws2.merge_cells(f"A{note_row}:G{note_row}")
    n = ws2.cell(row=note_row, column=1,
        value="* Overall Score = 40% Readability + 40% Topic Start Rate + 20% Speed")
    n.font = Font(name="Arial", italic=True, size=9, color="757575")
    n.alignment = LEFT

    for col, w in enumerate([16,18,14,12,16,14,14], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# Sidebar

with st.sidebar:
    st.header("⚙️ Configuration")
    st.subheader("🤖 Model Status")
    if check_backend():
        for m in get_model_status():
            icon = "🟢" if m["ready"] else "🔴"
            note = "ready" if m["ready"] else "not pulled"
            st.markdown(f"{icon} **{m['name']}** — {note}")
    else:
        st.error("Backend offline")
    st.markdown("---")
    st.info(f"⚡ CPU Mode: text capped at {MAX_CHARS:,} chars")

# Backend guard

if not check_backend():
    st.error("⚠️ Backend is not running. Start with:\n```\npython run_app.py\n```")
    st.stop()


# Tabs

tab_summarize, tab_matrix, tab_history = st.tabs(["📄 Summarize", "📊 Model Matrix", "📜 History"])

# Single Summarize

with tab_summarize:

    # Session state init
    if "queue_files" not in st.session_state:
        st.session_state["queue_files"] = []   # list of {name, text}
    if "job_history" not in st.session_state:
        st.session_state["job_history"] = []

    # Settings row 
    col1, col2 = st.columns(2)
    with col1:
        summary_type = st.selectbox(
            "1️⃣ Summary Type",
            SUMMARY_TYPES,
            format_func=lambda x: SUMMARY_LABELS[x],
            key="single_stype"
        )
    with col2:
        selected_model = st.selectbox("2️⃣ Model", MODELS, key="single_model")

    st.markdown("---")

    # File uploader + Add button
    st.markdown("**3️⃣ Add Documents to Queue**")
    col_up, col_add = st.columns([4, 1])

    with col_up:
        uploaded_file = st.file_uploader(
            "Select a file (PDF, DOCX, TXT)",
            type=["pdf", "docx", "txt"],
            key="uploader_single",
            label_visibility="collapsed"
        )
    with col_add:
        # Vertical spacer to align button with uploader
        st.markdown("<div style='margin-top:4px'></div>", unsafe_allow_html=True)
        add_clicked = st.button("➕ Add to Queue", use_container_width=True, key="btn_add")

    # Add file to queue list when Add is clicked
    if add_clicked:
        if uploaded_file is None:
            st.warning("Please select a file first.")
        else:
            already = [f["name"] for f in st.session_state["queue_files"]]
            if uploaded_file.name in already:
                st.warning(f"`{uploaded_file.name}` is already in the queue.")
            else:
                text = extract_text(uploaded_file)
                if text:
                    st.session_state["queue_files"].append({
                        "name": uploaded_file.name,
                        "text": text,
                        "chars": len(text),
                    })
                    st.success(f"✅ `{uploaded_file.name}` added to queue.")
                else:
                    st.error("Could not extract text from this file.")

    # Queue preview
    queue = st.session_state["queue_files"]

    if queue:
        st.markdown(f"**📋 Queue — {len(queue)} file(s)**")
        for i, f in enumerate(queue):
            col_name, col_chars, col_remove = st.columns([4, 2, 1])
            with col_name:
                st.markdown(f"📄 `{f['name']}`")
            with col_chars:
                st.caption(f"{f['chars']:,} chars")
            with col_remove:
                if st.button("✕", key=f"remove_{i}", help="Remove from queue"):
                    st.session_state["queue_files"].pop(i)
                    st.rerun()

        st.markdown("---")

        # Run all button
        if st.button(
            f"🚀 Run All ({len(queue)} file{'s' if len(queue) > 1 else ''})",
            type="primary", key="btn_run_all"
        ):
            progress = st.progress(0, text="Submitting jobs...")

            # Submit all jobs at once (non-blocking — this is where Job Queue shines)
            job_map = {}   # filename → job_id
            for f in queue:
                jid = submit_job(f["text"], selected_model, summary_type)
                if jid:
                    job_map[f["name"]] = jid
                    st.session_state["job_history"].append({
                        "job_id": jid,
                        "filename": f["name"],
                        "model": selected_model,
                        "summary_type": summary_type,
                    })

            st.info(f"📤 {len(job_map)} job(s) queued. Collecting results...")
            st.markdown("---")

            # Poll and display results one by one
            for idx, (filename, jid) in enumerate(job_map.items()):
                st.markdown(f"### 📄 `{filename}`")
                result = poll_until_done(jid, label=filename)

                if result:
                    st.markdown(
                        f'<div class="result-box">{result["summary"]}</div>',
                        unsafe_allow_html=True
                    )
                    st.markdown(f"""
                    <div class="status-box">
                        ⏱️ Time: <b>{result.get('time')}s</b> &nbsp;|&nbsp;
                        🤖 Model: <b>{result.get('model')}</b> &nbsp;|&nbsp;
                        📝 Words: <b>{len(result['summary'].split())}</b>
                    </div>""", unsafe_allow_html=True)
                    st.download_button(
                        f"💾 Download — {filename}",
                        data=result["summary"],
                        file_name=f"{Path(filename).stem}_{summary_type}.txt",
                        mime="text/plain",
                        key=f"dl_{jid}"
                    )
                else:
                    st.error(f"❌ `{filename}` failed.")

                progress.progress((idx + 1) / len(job_map),
                    text=f"Completed {idx+1}/{len(job_map)}")
                st.markdown("---")

            # Clear queue after run
            st.session_state["queue_files"] = []
            progress.progress(1.0, text="✅ All done!")

# Model Matrix Experiment

with tab_matrix:
    st.markdown("Run every **model × summary type** combination on one document and compare results.")
    st.markdown("---")

    # Settings row
    col_file, col_cat, col_models, col_types = st.columns([2, 1, 1, 1])

    with col_file:
        matrix_file = st.file_uploader(
            "Upload Test Document",
            type=["pdf", "docx", "txt"],
            key="uploader_matrix"
        )
    with col_cat:
        doc_category = st.selectbox(
            "Document Category",
            DOC_CATEGORIES,
            format_func=lambda x: CATEGORY_LABELS[x],
            key="matrix_category"
        )
    with col_models:
        selected_models = st.multiselect(
            "Models",
            MODELS,
            default=MODELS,
            key="matrix_models"
        )
    with col_types:
        selected_types = st.multiselect(
            "Summary Types",
            SUMMARY_TYPES,
            default=SUMMARY_TYPES,
            format_func=lambda x: SUMMARY_LABELS[x],
            key="matrix_types"
        )

    total_combos = len(selected_models) * len(selected_types)
    st.caption(f"Total combinations to run: **{total_combos}**")

    if matrix_file:
        matrix_text = extract_text(matrix_file)
        if not matrix_text:
            st.error("Could not extract text from file.")
        else:
            st.success(f"✅ `{matrix_file.name}` — {len(matrix_text):,} characters")

            if st.button("🚀 Run Matrix Experiment", type="primary", key="btn_matrix"):
                progress = st.progress(0, text="Submitting jobs...")
                completed_count = 0
                all_results = []

                # Submit all jobs at once
                job_map = {}
                for model in selected_models:
                    for stype in selected_types:
                        jid = submit_job(matrix_text, model, stype)
                        if jid:
                            job_map[(model, stype)] = jid

                st.info(f"📤 {len(job_map)} jobs queued. Collecting results...")

                # Live result grid
                st.subheader("Live Results")
                grid_cols = st.columns(len(selected_models))
                result_slots = {}
                for idx, model in enumerate(selected_models):
                    with grid_cols[idx]:
                        st.markdown(f"**🤖 {model}**")
                        for stype in selected_types:
                            result_slots[(model, stype)] = st.empty()
                            result_slots[(model, stype)].info(f"⏳ {SUMMARY_LABELS[stype]}...")

                # Poll and fill slots
                for (model, stype), jid in job_map.items():
                    result = poll_until_done(jid)
                    completed_count += 1
                    progress.progress(completed_count / total_combos,
                            text=f"Completed {completed_count}/{total_combos}")

                    slot = result_slots[(model, stype)]
                    if result:
                        metrics = evaluate_summary(result["summary"], result.get("time", 0))
                        all_results.append({
                            "doc_category":    doc_category,
                            "model":           model,
                            "summary_type":    stype,
                            "summary":         result["summary"],
                            "summary_preview": result["summary"][:200] + "...",
                            **metrics,
                        })
                        topic_icon = "✅" if metrics["starts_on_topic"] else "❌"
                        slot.success(
                            f"**{SUMMARY_LABELS[stype]}**  \n"
                            f"{topic_icon} Topic | ⏱ {metrics['time_taken']}s | "
                            f"📝 {metrics['word_count']}w | 📖 {metrics['readability_label']}"
                        )
                    else:
                        slot.error(f"❌ {SUMMARY_LABELS[stype]} failed")

                progress.progress(1.0, text="✅ Experiment complete!")

                # Results tables
                if all_results:
                    st.markdown("---")
                    st.subheader("Results")
                    df = pd.DataFrame(all_results)

                    r1, r2, r3 = st.columns(3)
                    with r1:
                        st.markdown("**⏱️ Time (seconds)**")
                        pivot_time = df.pivot_table(
                            index="summary_type", columns="model",
                            values="time_taken", aggfunc="mean"
                        ).round(1)
                        st.dataframe(pivot_time, use_container_width=True)

                    with r2:
                        st.markdown("**✅ Starts on Topic**")
                        pivot_topic = df.pivot_table(
                            index="summary_type", columns="model",
                            values="starts_on_topic", aggfunc="mean"
                        ).applymap(lambda x: "✅ Yes" if x >= 0.5 else "❌ No")
                        st.dataframe(pivot_topic, use_container_width=True)

                    with r3:
                        st.markdown("**📖 Readability**")
                        pivot_read = df.pivot_table(
                            index="summary_type", columns="model",
                            values="readability", aggfunc="mean"
                        ).round(1)
                        st.dataframe(pivot_read, use_container_width=True)

                    # Full summaries
                    with st.expander("📄 View All Summaries"):
                        for r in all_results:
                            st.markdown(
                                f'<div class="combo-header">🤖 {r["model"]} &nbsp;|&nbsp; '
                                f'{SUMMARY_LABELS[r["summary_type"]]} &nbsp;|&nbsp; '
                                f'{"✅ on topic" if r["starts_on_topic"] else "❌ off topic"} &nbsp;|&nbsp; '
                                f'⏱ {r["time_taken"]}s</div>',
                                unsafe_allow_html=True
                            )
                            st.markdown(
                                f'<div class="result-box">{r["summary"]}</div>',
                                unsafe_allow_html=True
                            )
                            st.markdown("")

                    # Save to session for export
                    st.session_state["matrix_results"] = all_results

    # Export
    if "matrix_results" in st.session_state:
        st.markdown("---")
        st.subheader("Export Results")
        col_xl, col_csv = st.columns(2)
        with col_xl:
            st.download_button(
                "📥 Download Excel (.xlsx)",
                data=build_excel(st.session_state["matrix_results"]),
                file_name="docsum_matrix_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_csv:
            df_exp = pd.DataFrame(st.session_state["matrix_results"])
            st.download_button(
                "📥 Download CSV",
                data=df_exp.drop(columns=["summary"], errors="ignore").to_csv(index=False).encode(),
                file_name="docsum_matrix_results.csv",
                mime="text/csv",
                use_container_width=True,
            )

# History

with tab_history:
    history = st.session_state.get("job_history", [])
    if not history:
        st.info("No jobs yet in this session.")
    else:
        for entry in reversed(history):
            job = poll_job_once(entry["job_id"])
            if not job:
                continue
            status = job["status"]
            col_a, col_b, col_c, col_d = st.columns([3, 1, 1, 1])
            with col_a:
                st.markdown(f"**{entry['filename']}**")
                st.caption(f"`{entry['job_id'][:12]}...`")
            with col_b:
                st.markdown(f"`{entry['model']}`")
            with col_c:
                st.markdown(f"_{entry['summary_type']}_")
            with col_d:
                color = {"completed": "🟢", "failed": "🔴", "processing": "🟡"}.get(status, "⚪")
                st.markdown(f"{color} {status}")
            if status == "completed":
                with st.expander("View Summary"):
                    st.markdown(
                        f'<div class="result-box">{job.get("summary","")}</div>',
                        unsafe_allow_html=True
                    )
            st.markdown("---")
